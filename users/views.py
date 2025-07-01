from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, reverse, get_object_or_404
from django.contrib import messages
from datetime import datetime

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_workbook
from django.http import HttpResponse

from .forms import Step1Form, Step2Form, AddLeaderForm, AddParticipantForm, ProjectInfoForm
from .models import Application, Leader, Participant


def step1_view(request):
    if request.method == 'POST':
        form = Step1Form(request.POST)
        if form.is_valid():
            request.session['step1_data'] = form.cleaned_data
            return redirect('users:step2')
    else:
        form = Step1Form(initial=request.session.get('step1_data'))

    return render(request, 'users/step1.html', {'form': form})


def step2_view(request):
    """Шаг 2: Добавление руководителей."""
    step1_data = request.session.get('step1_data')
    if not step1_data:
        return redirect(reverse('users:step1'))

    leaders = request.session.get('leaders', [])
    add_leader_form = AddLeaderForm()

    if request.method == 'POST':
        if 'add_leader' in request.POST:
            add_leader_form = AddLeaderForm(request.POST)
            if add_leader_form.is_valid():
                leader_data = add_leader_form.cleaned_data
                leaders.append(leader_data)
                request.session['leaders'] = leaders
                add_leader_form = AddLeaderForm()

        elif 'remove_leader' in request.POST:
            leader_index = int(request.POST['remove_leader'])
            if 0 <= leader_index < len(leaders):
                leaders.pop(leader_index)
                request.session['leaders'] = leaders

        elif 'complete_step2' in request.POST:
            if leaders:
                request.session['leaders'] = leaders
                return redirect(reverse('users:step_project'))
            else:
                messages.error(request, "Пожалуйста, добавьте хотя бы одного руководителя.")

    context = {
        'leaders': leaders,
        'add_leader_form': add_leader_form,
    }
    return render(request, 'users/step2.html', context)


def project_info_view(request):
    """Шаг с информацией о проекте (перед добавлением участников)."""
    step1_data = request.session.get('step1_data')
    if not step1_data:
        return redirect('users:step1')

    project_info = request.session.get('project_info', {})

    if request.method == "POST":
        form = ProjectInfoForm(request.POST, initial=project_info)
        if form.is_valid():
            request.session['project_info'] = form.cleaned_data
            return redirect('users:step3')
    else:
        form = ProjectInfoForm(initial=project_info)

    return render(request, "users/step_project.html", {"form": form})


def step3_view(request):
    participants = request.session.get('participants', [])
    project_info = request.session.get('project_info', {})
    nomination = project_info.get('nomination')

    if request.method == 'POST':
        if 'add_participant' in request.POST:
            form = AddParticipantForm(request.POST, nomination=nomination)
            if form.is_valid():
                participant_data = {
                    'full_name': form.cleaned_data['full_name'],
                    'birth_date': str(form.cleaned_data['birth_date']),
                    'participation_type': form.cleaned_data['participation_type'],
                }

                participation_type = form.cleaned_data['participation_type']

                if participation_type == 'school':
                    participant_data['school_name'] = form.cleaned_data['school_name']
                    participant_data['grade'] = form.cleaned_data['grade']

                elif participation_type == 'college':
                    participant_data['college_name'] = form.cleaned_data['college_name']
                    participant_data['course'] = form.cleaned_data['course']

                elif participation_type == 'additional':
                    participant_data['additional_education_name'] = form.cleaned_data['additional_education_name']
                    participant_data['school_name'] = form.cleaned_data['school_name']
                    participant_data['grade'] = form.cleaned_data['grade']

                elif participation_type == 'movement':
                    participant_data['movement_type'] = form.cleaned_data['movement_type']

                elif participation_type == 'family':
                    participant_data['family_name'] = form.cleaned_data['family_name']

                elif participation_type == "kindergarten":
                    participant_data["kindergarten_name"] = form.cleaned_data["kindergarten_name"]

                elif participation_type == "family_education":
                    participant_data["family_education_surname"] = form.cleaned_data["family_education_surname"]

                participants.append(participant_data)
                request.session['participants'] = participants
                request.session.modified = True

                if participation_type == 'family':
                    # Создаём новую форму с теми же participation_type и family_name,
                    # но очищенными full_name и birth_date
                    initial_data = {
                        'participation_type': participation_type,
                        'family_name': form.cleaned_data['family_name'],
                    }
                    form = AddParticipantForm(initial=initial_data, nomination=nomination)
                    # Отрисовываем страницу без redirect, чтобы поля сохранились
                    return render(request, 'users/step3.html', {
                        'add_participant_form': form,
                        'participants': participants,
                    })

                else:
                    # Для остальных делаем redirect (обновляем страницу с пустой формой)
                    return redirect('users:step3')

        elif 'remove_participant' in request.POST:
            index = int(request.POST.get('remove_participant'))
            if 0 <= index < len(participants):
                participants.pop(index)
                request.session['participants'] = participants
                request.session.modified = True
                return redirect('users:step3')

        elif 'complete_step3' in request.POST:
            return redirect('users:complete_step4')

    else:
        form = AddParticipantForm(nomination=nomination)

    return render(request, 'users/step3.html', {
        'add_participant_form': form,
        'participants': participants,
    })



def complete_step4(request):
    """Финальный шаг: ввод комментария, сохранение в БД и завершение регистрации."""
    step1_data = request.session.get('step1_data')
    project_info = request.session.get('project_info')
    leaders = request.session.get('leaders')
    participants = request.session.get('participants')

    if not all([step1_data, project_info, leaders, participants]):
        if step1_data is None:
            return redirect('users:step1')
        elif not project_info:
            return redirect('users:step_project')
        elif not leaders:
            return redirect('users:step2')
        else:
            return redirect('users:step3')

    if request.method == 'POST':
        form = Step2Form(request.POST)  # В форме только комментарий
        if form.is_valid():
            comment = form.cleaned_data['comment']
            try:
                application = Application.objects.create(
                    region=step1_data['region'],
                    city=step1_data['city'],
                    organization_name=step1_data['organization_name'],
                    postal_address=step1_data['postal_address'],
                    phone_number=step1_data['phone_number'],
                    email=step1_data['email'],
                    website=step1_data['website'],
                    project_title=project_info.get('project_title'),
                    nomination=project_info.get('nomination'),
                    comment=comment,
                )

                for leader_data in leaders:
                    leader = Leader.objects.filter(
                        full_name=leader_data['full_name'],
                        position=leader_data.get('position', ''),
                        academic_title=leader_data.get('academic_title', ''),
                        workplace=leader_data.get('workplace', ''),
                        mobile_phone=leader_data['mobile_phone'],
                        other_contact=leader_data.get('other_contact', ''),
                    ).first()

                    if leader is None:
                        leader = Leader.objects.create(
                            full_name=leader_data['full_name'],
                            position=leader_data.get('position', ''),
                            academic_title=leader_data.get('academic_title', ''),
                            workplace=leader_data.get('workplace', ''),
                            mobile_phone=leader_data['mobile_phone'],
                            other_contact=leader_data.get('other_contact', ''),
                        )
                    application.leaders.add(leader)

                for participant_data in participants:
                    birth_date_dt = datetime.strptime(participant_data['birth_date'], '%Y-%m-%d').date()

                    participant_kwargs = {
                        "full_name": participant_data['full_name'],
                        "birth_date": birth_date_dt,
                        "participation_type": participant_data['participation_type'],
                    }

                    if participant_data['participation_type'] == "school":
                        participant_kwargs["school_name"] = participant_data.get("school_name", "")
                        participant_kwargs["grade"] = participant_data.get("grade", "")
                    elif participant_data['participation_type'] == "college":
                        participant_kwargs["college_name"] = participant_data.get("college_name", "")
                        participant_kwargs["course"] = participant_data.get("course", "")
                    elif participant_data['participation_type'] == "additional":
                        participant_kwargs["additional_education_name"] = participant_data.get("additional_education_name", "")
                        participant_kwargs["school_name"] = participant_data.get("school_name", "")
                        participant_kwargs["grade"] = participant_data.get("grade", "")
                    elif participant_data['participation_type'] == "movement":
                        participant_kwargs["movement_type"] = participant_data.get("movement_type", "")
                    elif participant_data['participation_type'] == "family":

                        participant_kwargs["family_name"] = participant_data.get("family_name", "")
                    elif participant_data['participation_type'] == "kindergarten":
                        participant_kwargs["kindergarten_name"] = participant_data.get("kindergarten_name", "")
                    elif participant_data['participation_type'] == "family_education":
                        participant_kwargs["family_education_surname"] = participant_data.get("family_education_surname", "")

                    participant = Participant.objects.create(**participant_kwargs)
                    application.participants.add(participant)
                registration_number = application.registration_number

                request.session.flush()
                request.session['registration_number'] = registration_number


                return redirect('users:success')

            except Exception as e:
                return render(request, 'users/error.html', {'error_message': str(e)})
    else:
        form = Step2Form()
    return render(request, 'users/step4.html', {'form': form})



def success_view(request):
    registration_number = request.session.get('registration_number')
    return render(request, 'users/success.html', {'registration_number': registration_number})


def error_view(request):
    return render(request, 'users/error.html')


@staff_member_required
def export_applications_detailed_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = [
        'Регистрационный номер',
        'Тип записи',
        'ФИО',
        'Дата рождения',
        'Тип участия',
        'Название семейного коллектива',
        'Школа',
        'Класс',
        'Колледж/ВУЗ',
        'Курс',
        'Учреждение ДО',
        'Тип движения',
        'Детский сад',
        'Семейное воспитание (фамилия)',
        'Должность',
        'Ученое звание',
        'Место работы',
        'Телефон',
        'Другой контакт',
        'Тема проекта',
        'Регион',
        'Город',
        'Название организации',
        'Почтовый адрес',
        'Телефон организации',
        'Email',
        'Сайт',
        'Комментарий',
    ]
    ws.append(headers)

    for app in Application.objects.all():
        reg_num = app.registration_number or ''
        project_title = app.project_title or ''
        region_display = app.region or ''

        # Участники
        for participant in app.participants.all():
            ws.append([
                reg_num,
                'Участник',
                participant.full_name,
                participant.birth_date.strftime('%d.%m.%Y') if participant.birth_date else '',
                participant.get_participation_type_display() if participant.participation_type else '',
                participant.family_name or '',
                participant.school_name or '',
                participant.grade or '',
                participant.college_name or '',
                participant.course or '',
                participant.additional_education_name or '',
                participant.get_movement_type_display() or '',
                participant.kindergarten_name or '',
                participant.family_education_surname or '',
                '',  # Должность у участника отсутствует
                '',  # Ученое звание отсутствует
                '',  # Место работы отсутствует
                '',  # Телефон у участника отсутствует
                '',  # Другой контакт у участника отсутствует
                project_title,
                region_display,
                app.city or '',
                app.organization_name or '',
                app.postal_address or '',
                app.phone_number or '',
                app.email or '',
                app.website or '',
                app.comment or '',
            ])

        # Руководители
        for leader in app.leaders.all():
            ws.append([
                reg_num,
                'Руководитель',
                leader.full_name,
                '',  # Дата рождения отсутствует
                '',  # Тип участия отсутствует
                '',  # Семейный коллектив отсутствует
                '',  # Школа отсутствует
                '',  # Класс отсутствует
                '',  # Колледж отсутствует
                '',  # Курс отсутствует
                '',  # Учреждение ДО отсутствует
                '',  # Тип движения отсутствует
                '',  # Детский сад отсутствует
                '',  # Фамилия для семейного коллектива отсутствует
                leader.position or '',
                leader.academic_title or '',
                leader.workplace or '',
                leader.mobile_phone or '',
                leader.other_contact or '',
                project_title,
                region_display,
                app.city or '',
                app.organization_name or '',
                app.postal_address or '',
                app.phone_number or '',
                app.email or '',
                app.website or '',
                app.comment or '',
            ])

    # Автоподгонка ширины столбцов
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="applications_full.xlsx"'
    wb.save(response)
    return response


def applications_list(request):
    return render(request, 'users/applications_list.html')
